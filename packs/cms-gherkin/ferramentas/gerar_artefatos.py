from __future__ import annotations

import argparse
import json
import re
import textwrap
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "FONTES-ESTRUTURADAS"
ANALYSIS_DIR = ROOT / "ANALISES"
BDD_DIR = ROOT / "Processados_BDD"
FEATURE_DIR = ROOT / "xray_export"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
WRAP_TOP = Alignment(vertical="top", wrap_text=True)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Gera analise, planilha e feature a partir da fonte estruturada."
    )
    parser.add_argument(
        "--fonte",
        type=Path,
        help="Arquivo JSON em FONTES-ESTRUTURADAS. Se omitido, processa todos.",
    )
    args = parser.parse_args()

    sources = (
        [resolve_source(args.fonte)]
        if args.fonte
        else sorted(path for path in SOURCE_DIR.glob("*.json") if not path.name.startswith("MODELO-"))
    )
    if not sources:
        raise SystemExit("Nenhuma fonte estruturada encontrada.")

    for source_path in sources:
        data = load_source(source_path)
        slug = data["slug"]

        analysis_path = ANALYSIS_DIR / f"{slug}-ANALISE-DETALHADA.md"
        spreadsheet_path = BDD_DIR / f"{slug}.xlsx"
        feature_path = FEATURE_DIR / f"{slug}.feature"

        analysis_path.write_text(render_analysis(data), encoding="utf-8")
        build_workbook(data).save(spreadsheet_path)
        feature_path.write_text(render_feature(data), encoding="utf-8")

        print(f"[ok] Fonte: {source_path.relative_to(ROOT)}")
        print(f"     Analise: {analysis_path.relative_to(ROOT)}")
        print(f"     Planilha: {spreadsheet_path.relative_to(ROOT)}")
        print(f"     Feature: {feature_path.relative_to(ROOT)}")

    return 0


def resolve_source(source_path: Path | None) -> Path:
    if source_path is None:
        raise SystemExit("Informe uma fonte valida.")

    candidates = [source_path]
    if not source_path.is_absolute():
        candidates.extend(
            [
                Path.cwd() / source_path,
                ROOT / source_path,
                SOURCE_DIR / source_path.name,
            ]
        )

    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()

    raise SystemExit(f"Fonte nao encontrada: {source_path}")


def load_source(path: Path) -> dict:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SystemExit(f"JSON invalido em {path}: {exc}") from exc

    validate_source(data, path)
    return data


def validate_source(data: dict, path: Path) -> None:
    required_fields = [
        "slug",
        "title",
        "summary",
        "business_problem",
        "functional_goal",
        "main_context",
        "dependencies",
        "risks",
        "test_data",
        "acceptance_criteria",
        "business_rules",
        "scenario_strategy",
        "modeling_decisions",
        "feature_description",
        "scenarios",
        "conclusion",
    ]

    missing = [field for field in required_fields if field not in data]
    if missing:
        raise SystemExit(f"{path.name} esta sem os campos obrigatorios: {', '.join(missing)}")

    scenario_ids: set[int] = set()
    ac_ids = {normalize_ac_id(item["id"]) for item in data["acceptance_criteria"]}
    covered_acs: set[str] = set()

    for item in data["acceptance_criteria"]:
        if "id" not in item or "text" not in item:
            raise SystemExit("Cada criterio de aceite precisa de 'id' e 'text'.")

    for item in data["business_rules"]:
        if "id" not in item or "title" not in item:
            raise SystemExit("Cada regra de negocio precisa de 'id' e 'title'.")

    for scenario in data["scenarios"]:
        scenario_required = ("id", "name", "bdd", "acs", "explanation")
        for field in scenario_required:
            if field not in scenario:
                raise SystemExit(
                    f"Cenario incompleto em {path.name}: faltou '{field}' no item {scenario!r}"
                )

        scenario_id = scenario["id"]
        if scenario_id in scenario_ids:
            raise SystemExit(f"ID de cenario duplicado em {path.name}: {scenario_id}")
        scenario_ids.add(scenario_id)

        if not has_required_steps(scenario["bdd"]):
            raise SystemExit(
                f"O BDD do cenario {scenario_id} precisa conter Dado, Quando e Então/Entao."
            )

        scenario_acs = {normalize_ac_id(ac) for ac in scenario["acs"]}
        unknown_acs = scenario_acs - ac_ids
        if unknown_acs:
            raise SystemExit(
                f"Cenario {scenario_id} referencia AC inexistente: {', '.join(sorted(unknown_acs))}"
            )
        covered_acs.update(scenario_acs)

    missing_coverage = ac_ids - covered_acs
    if missing_coverage:
        raise SystemExit(
            "Cobertura incompleta. ACs sem cenario: "
            + ", ".join(sorted(missing_coverage))
        )


def has_required_steps(bdd: str) -> bool:
    lowered = collapse_space(bdd).casefold()
    return all(word in lowered for word in ("dado", "quando", "então")) or all(
        word in lowered for word in ("dado", "quando", "entao")
    )


def render_analysis(data: dict) -> str:
    slug = data["slug"]
    lines = [f"# {slug} - Analise Detalhada", ""]

    lines.extend(["## 1. Identificacao", ""])
    lines.append(f"- Projeto: `{slug}`")
    lines.append(f"- Titulo: `{data['title']}`")
    lines.append(f"- Versao: `{data.get('version', 'v1')}`")
    source_documents = data.get("source_documents", [])
    if source_documents:
        lines.append("- Artefatos associados:")
        for item in source_documents:
            lines.append(f"  - `{item}`")
    lines.append(f"  - `ANALISES/{slug}-ANALISE-DETALHADA.md`")
    lines.append(f"  - `Processados_BDD/{slug}.xlsx`")
    lines.append(f"  - `xray_export/{slug}.feature`")
    lines.append("")

    lines.extend(["## 2. Leitura funcional da US", "", "### Problema de negocio", ""])
    lines.append(data["business_problem"])
    lines.extend(["", "### Objetivo funcional", ""])
    lines.append(data["functional_goal"])
    lines.extend(["", "### Contexto principal extraido da US", ""])
    lines.extend(format_bullets(data["main_context"]))
    lines.append("")

    lines.extend(["## 3. Regras de negocio extraidas", ""])
    for rule in data["business_rules"]:
        lines.append(f"### {rule['id']} - {rule['title']}")
        lines.append("")
        description = rule.get("description")
        if description:
            lines.append(description)
            lines.append("")
        items = rule.get("items", [])
        if items:
            lines.extend(format_bullets(items))
            lines.append("")

    lines.extend(["## 4. ACs oficiais consolidados", ""])
    for item in data["acceptance_criteria"]:
        lines.append(f"- `{item['id']}`: {item['text']}")
    lines.append("")

    lines.extend(["## 5. Raciocinio utilizado para criar os cenarios", "", "### Estrategia de decomposicao", ""])
    lines.append(data["scenario_strategy"])
    lines.append("")
    for scenario in data["scenarios"]:
        lines.append(f"### Cenario {scenario['id']} - {scenario['name']}")
        lines.append("")
        lines.append(scenario["explanation"])
        lines.append("")

    lines.extend(["## 6. Decisoes de modelagem", ""])
    lines.extend(format_bullets(data["modeling_decisions"]))
    lines.append("")

    lines.extend(["## 7. Cobertura final", ""])
    lines.append(f"- total de ACs oficiais: `{len(data['acceptance_criteria'])}`")
    lines.append(f"- total de cenarios finais: `{len(data['scenarios'])}`")
    lines.append("- cobertura declarada: `100%`")
    lines.append("")

    lines.extend(["## 8. Conclusao", "", data["conclusion"], ""])
    return "\n".join(lines)


def build_workbook(data: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(data["slug"])

    set_widths(ws)
    row = 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    title_cell = ws.cell(row=row, column=2, value=f"{data['slug']} - {data['title']}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = WRAP_TOP
    row += 2

    row = write_label_value(ws, row, "Resumo", data["summary"])
    row = write_label_value(ws, row, "Dependencias principais", join_lines(data["dependencies"]))
    row = write_label_value(ws, row, "Riscos principais", join_lines(data["risks"]))
    row = write_label_value(ws, row, "Massas de teste sugeridas", join_lines(data["test_data"]))
    row = write_label_value(
        ws,
        row,
        "Criterios de Aceite",
        join_lines([f"(OK) - {item['text']}" for item in data["acceptance_criteria"]]),
    )

    row += 1
    row = write_table_header(ws, row, ["Tema", "Pergunta / ponto de atencao"])
    for item in data.get("attention_points", []):
        ws.cell(row=row, column=2, value=item["theme"]).alignment = WRAP_TOP
        ws.cell(row=row, column=3, value=item["question"]).alignment = WRAP_TOP
        row += 1

    row += 2
    row = write_table_header(ws, row, ["# AC", "Criterios de Aceite Oficiais"])
    for item in data["acceptance_criteria"]:
        ws.cell(row=row, column=2, value=item["id"]).alignment = WRAP_TOP
        ws.cell(row=row, column=3, value=item["text"]).alignment = WRAP_TOP
        row += 1

    row += 2
    row = write_table_header(ws, row, ["#", "Cenario", "BDD", "AC Relacionado"])
    for scenario in data["scenarios"]:
        ws.cell(row=row, column=2, value=scenario["id"]).alignment = WRAP_TOP
        ws.cell(row=row, column=3, value=scenario["name"]).alignment = WRAP_TOP
        ws.cell(row=row, column=4, value=collapse_space(scenario["bdd"])).alignment = WRAP_TOP
        ws.cell(row=row, column=5, value=", ".join(scenario["acs"])).alignment = WRAP_TOP
        row += 1

    return wb


def render_feature(data: dict) -> str:
    generated_at = data.get("generated_at", str(date.today()))
    feature_tags = unique_tags([data["slug"], *data.get("feature_tags", [])])
    description_lines = wrap_feature_description(data["feature_description"])

    lines = [
        "# language: pt",
        "# encoding: utf-8",
        f"# {data['slug']} - {data['title']}",
        f"# Versao: {data.get('version', 'v1')}",
        f"# Total de Cenarios: {len(data['scenarios'])}",
        f"# Gerado em: {generated_at}",
        "",
        " ".join(f"@{tag}" for tag in feature_tags),
        f"Funcionalidade: {data['slug']} - {data['title']}",
    ]
    lines.extend(f"  {line}" for line in description_lines)

    for scenario in data["scenarios"]:
        lines.append("")
        tags = [f"@CT-{scenario['id']}"] + [f"@{to_ac_tag(ac)}" for ac in scenario["acs"]]
        lines.append(f"  {' '.join(tags)}")
        lines.append(f"  Cenário: {scenario['name']}")
        for step in split_bdd_steps(scenario["bdd"]):
            lines.append(f"    {step}")

    lines.append("")
    return "\n".join(lines)


def wrap_feature_description(text: str) -> list[str]:
    return textwrap.wrap(text, width=66) or [text]


def split_bdd_steps(bdd: str) -> list[str]:
    pattern = r"\b(Dado|Quando|Então|Entao|E|Mas)\b"
    matches = list(re.finditer(pattern, bdd))
    if not matches:
        return [collapse_space(bdd)]

    steps: list[str] = []
    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(bdd)
        steps.append(collapse_space(bdd[start:end]))
    return steps


def write_label_value(ws, row: int, label: str, value: str) -> int:
    label_cell = ws.cell(row=row, column=2, value=label)
    label_cell.font = Font(bold=True)
    label_cell.fill = SECTION_FILL
    label_cell.alignment = WRAP_TOP

    value_cell = ws.cell(row=row, column=3, value=value)
    value_cell.alignment = WRAP_TOP
    return row + 1


def write_table_header(ws, row: int, headers: list[str]) -> int:
    for offset, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=offset, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_TOP
    return row + 1


def set_widths(ws) -> None:
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 85
    ws.column_dimensions["E"].width = 22


def join_lines(items: list[str]) -> str:
    return "\n".join(items)


def format_bullets(items: list[str]) -> list[str]:
    return [f"- {item}" for item in items]


def collapse_space(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def normalize_ac_id(ac_id: str) -> str:
    match = re.search(r"(\d+)", ac_id)
    if not match:
        raise SystemExit(f"AC invalido: {ac_id}")
    return f"AC {match.group(1)}"


def to_ac_tag(ac_id: str) -> str:
    return normalize_ac_id(ac_id).replace(" ", "-")


def unique_tags(items: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for item in items:
        cleaned = item.strip().replace(" ", "-")
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            ordered.append(cleaned)
    return ordered


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "-", title)
    return cleaned[:31]


if __name__ == "__main__":
    raise SystemExit(main())
